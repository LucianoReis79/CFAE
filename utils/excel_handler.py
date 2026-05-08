import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os
import re
from pathlib import Path
from datetime import datetime

def obter_abas(file):
    file.seek(0)
    wb = load_workbook(file, read_only=True)
    return wb.sheetnames

def obter_colunas(file, aba):
    file.seek(0)
    df = pd.read_excel(file, sheet_name=aba, nrows=1)
    return df.columns.tolist()

def obter_valores_unicos(file, aba, coluna):
    file.seek(0)
    df = pd.read_excel(file, sheet_name=aba, usecols=[coluna])
    return df[coluna].dropna().unique().tolist()

def limpar_nome(nome):
    return re.sub(r'[\\/*?:"<>|]', "", str(nome))

def verificar_pasta_downloads():
    return str(Path.home() / "Downloads")

def gerar_arquivos_filtrados(file, aba_nome, col_filtro, lista_valores, destino, pbar, status_txt):
    erros = []
    total = len(lista_valores)
    mes_ano = datetime.now().strftime("%m_%Y")

    for idx, valor in enumerate(lista_valores):
        try:
            status_txt.text(f"Processando: {valor}")
            file.seek(0)
            
            # Carregar o workbook. data_only=False preserva formatação.
            wb = load_workbook(file, data_only=False)
            
            # Remover outras abas para manter o arquivo leve
            for sheet in wb.sheetnames:
                if sheet != aba_nome:
                    wb.remove(wb[sheet])
            
            ws = wb[aba_nome]

            # Solução para erro de Tabelas (table1.xml)
            table_names = list(ws.tables.keys())
            for table_name in table_names:
                del ws.tables[table_name]
            
            # --- BUSCA ROBUSTA DA COLUNA ---
            col_idx = None
            # Normalizamos o nome da coluna procurada
            alvo = str(col_filtro).strip().upper()
            
            # Iteramos apenas a primeira linha (cabeçalho)
            for cell in ws[5]: 
                if cell.value:
                    # Limpamos o valor da célula de quebras de linha (\n) e espaços
                    valor_cabecalho = str(cell.value).replace('\n', ' ').strip().upper()
                    if valor_cabecalho == alvo:
                        col_idx = cell.column
                        break
            
            if col_idx is None:
                # Se não achar, tenta uma busca parcial como último recurso
                for cell in ws[5]:
                    if cell.value and alvo in str(cell.value).upper():
                        col_idx = cell.column
                        break

            if col_idx is None:
                raise ValueError(f"A coluna '{col_filtro}' não foi localizada na linha 1 da aba '{aba_nome}'.")

            # Deletar linhas de baixo para cima
            for row_num in range(ws.max_row, 1, -1):
                celula_valor = ws.cell(row=row_num, column=col_idx).value
                
                # Comparação robusta de valores
                if str(celula_valor).strip() != str(valor).strip():
                    ws.delete_rows(row_num)

            # Salvar o arquivo
            valor_limpo = limpar_nome(valor)
            nome_arq = f"{limpar_nome(col_filtro)}_{valor_limpo}_{mes_ano}.xlsx"
            wb.save(os.path.join(destino, nome_arq))
            
        except Exception as e:
            erros.append(f"{valor}: {str(e)}")
        
        pbar.progress((idx + 1) / total)

    return True, erros