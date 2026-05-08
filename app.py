import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import tempfile
import os
import shutil
from pathlib import Path
from datetime import datetime

st.set_page_config(
    page_title="Gerador de Arquivos",
    layout="wide"
)

st.title("📁 Gerador de Arquivos Filtrados")

arquivo = st.file_uploader(
    "Selecione a planilha Excel",
    type=["xlsx", "xlsm"]
)

if arquivo:

    try:

        # =========================
        # GARANTIR LEITURA DO BUFFER
        # =========================

        arquivo.seek(0)

        xls = pd.ExcelFile(arquivo)

        abas = xls.sheet_names

        # =========================
        # SELEÇÃO DA ABA
        # =========================

        aba_escolhida = st.selectbox(
            "Selecione a aba",
            abas,
            index=abas.index("BASE_NÚCLEO")
            if "BASE_NÚCLEO" in abas else 0
        )

        # =========================
        # LER DADOS
        # =========================

        arquivo.seek(0)

        df = pd.read_excel(
            arquivo,
            sheet_name=aba_escolhida
        )

        # =========================
        # COLUNA DE FILTRO
        # =========================

        coluna_filtro = st.selectbox(
            "Selecione a coluna para filtro",
            df.columns
        )

        valores = sorted(
            df[coluna_filtro]
            .dropna()
            .astype(str)
            .str.strip()
            .unique()
        )

        # =========================
        # MODO
        # =========================

        modo = st.radio(
            "Modo de geração",
            [
                "Gerar todos",
                "Gerar apenas selecionados"
            ]
        )

        if modo == "Gerar apenas selecionados":

            selecionados = st.multiselect(
                "Selecione os valores",
                valores
            )

        else:

            selecionados = valores

        # =========================
        # NOME PADRÃO
        # =========================

        nome_padrao = st.text_input(
            "Nome padrão dos arquivos",
            value="BASE_NÚCLEO"
        )

        # =========================
        # BOTÃO
        # =========================

        if st.button("🚀 Gerar Arquivos"):

            try:

                if len(selecionados) == 0:

                    st.warning("Nenhum item selecionado.")
                    st.stop()

                pasta_downloads = str(
                    Path.home() / "Downloads"
                )

                progresso = st.progress(0)

                total = len(selecionados)

                # =========================
                # EXTENSÃO ORIGINAL
                # =========================

                extensao = (
                    ".xlsm"
                    if arquivo.name.endswith(".xlsm")
                    else ".xlsx"
                )

                # =========================
                # CRIAR ARQUIVO TEMPORÁRIO
                # =========================

                with tempfile.NamedTemporaryFile(
                    delete=False,
                    suffix=extensao
                ) as tmp:

                    arquivo.seek(0)

                    shutil.copyfileobj(
                        arquivo,
                        tmp
                    )

                    caminho_temp = tmp.name

                # =========================
                # LOOP
                # =========================

                for i, valor in enumerate(selecionados):

                    wb = load_workbook(
                        filename=caminho_temp,
                        keep_vba=arquivo.name.endswith(".xlsm")
                    )

                    ws = wb[aba_escolhida]

                    # =========================
                    # MAPEAR COLUNAS
                    # =========================

                    colunas = {}

                    for idx, cell in enumerate(
                        ws[1],
                        start=1
                    ):

                        nome_coluna = str(
                            cell.value
                        ).strip()

                        colunas[nome_coluna] = idx

                    coluna_idx = colunas[coluna_filtro]

                    # =========================
                    # IDENTIFICAR LINHAS
                    # =========================

                    linhas_para_remover = []

                    for row in range(
                        2,
                        ws.max_row + 1
                    ):

                        valor_celula = ws.cell(
                            row=row,
                            column=coluna_idx
                        ).value

                        valor_celula = str(
                            valor_celula
                        ).strip()

                        if valor_celula != str(valor).strip():

                            linhas_para_remover.append(row)

                    # =========================
                    # REMOVER LINHAS
                    # =========================

                    for row in reversed(
                        linhas_para_remover
                    ):

                        ws.delete_rows(row)

                    # =========================
                    # REMOVER OUTRAS ABAS
                    # =========================

                    for nome_aba in wb.sheetnames:

                        if nome_aba != aba_escolhida:

                            del wb[nome_aba]

                    # =========================
                    # NOME DO ARQUIVO
                    # =========================

                    mes_ano = datetime.now().strftime(
                        "%m_%Y"
                    )

                    nome_arquivo = (
                        f"{nome_padrao}_"
                        f"{valor}_"
                        f"{mes_ano}"
                        f"{extensao}"
                    )

                    caracteres_invalidos = [
                        "\\",
                        "/",
                        ":",
                        "*",
                        "?",
                        "\"",
                        "<",
                        ">",
                        "|"
                    ]

                    for c in caracteres_invalidos:

                        nome_arquivo = nome_arquivo.replace(
                            c,
                            "_"
                        )

                    caminho_saida = os.path.join(
                        pasta_downloads,
                        nome_arquivo
                    )

                    # =========================
                    # FORÇAR RECÁLCULO
                    # =========================

                    wb.calculation.fullCalcOnLoad = True
                    wb.calculation.forceFullCalc = True

                    # =========================
                    # SALVAR
                    # =========================

                    wb.save(caminho_saida)

                    progresso.progress(
                        (i + 1) / total
                    )

                st.success(
                    f"✅ {total} arquivos gerados com sucesso!"
                )

                st.info(
                    f"Arquivos salvos em:\n{pasta_downloads}"
                )

            except Exception as e:

                st.error(
                    f"Erro ao gerar arquivos:\n\n{e}"
                )

    except Exception as e:

        st.error(
            f"Erro ao carregar arquivo:\n\n{e}"
        )