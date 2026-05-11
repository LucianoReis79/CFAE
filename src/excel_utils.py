import os
import shutil
import tempfile
from copy import copy
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook

from src.file_utils import clean_filename


def load_workbook_data(file_path):

    workbook = load_workbook(
        file_path,
        data_only=False
    )

    return workbook


def get_sheet_names(workbook):

    return workbook.sheetnames


def get_columns(
    workbook,
    sheet_name,
    return_unique=False,
    filter_column=None
):

    sheet = workbook[sheet_name]

    headers = [cell.value for cell in sheet[1]]

    if not return_unique:
        return headers

    column_index = headers.index(filter_column) + 1

    values = set()

    for row in sheet.iter_rows(
        min_row=2,
        values_only=True
    ):
        value = row[column_index - 1]

        if value is not None:
            values.add(str(value))

    return sorted(values)


def copy_cell_style(source, target):

    if source.has_style:
        target._style = copy(source._style)

    if source.number_format:
        target.number_format = source.number_format

    if source.font:
        target.font = copy(source.font)

    if source.fill:
        target.fill = copy(source.fill)

    if source.border:
        target.border = copy(source.border)

    if source.alignment:
        target.alignment = copy(source.alignment)

    if source.protection:
        target.protection = copy(source.protection)


def convert_formulas_to_values(sheet):

    for row in sheet.iter_rows():
        for cell in row:
            cell.value = cell.value


def remove_excel_tables(sheet):

    if sheet.tables:
        sheet.tables.clear()


def remove_non_matching_rows(
    sheet,
    filter_column,
    filter_value
):

    headers = [cell.value for cell in sheet[1]]

    col_index = headers.index(filter_column) + 1

    rows_to_delete = []

    for row in range(2, sheet.max_row + 1):

        cell_value = sheet.cell(
            row=row,
            column=col_index
        ).value

        if str(cell_value) != str(filter_value):
            rows_to_delete.append(row)

    for row in reversed(rows_to_delete):
        sheet.delete_rows(row, 1)


def keep_only_selected_sheet(
    workbook,
    selected_sheet
):

    for sheet in workbook.sheetnames:

        if sheet != selected_sheet:
            std = workbook[sheet]
            workbook.remove(std)


def preserve_column_widths(
    original_sheet,
    target_sheet
):

    for col_letter, dim in original_sheet.column_dimensions.items():
        target_sheet.column_dimensions[col_letter].width = dim.width


def preserve_merged_cells(
    original_sheet,
    target_sheet
):

    for merged_range in original_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merged_range))


def generate_filtered_files(
    source_file,
    sheet_name,
    filter_column,
    values,
    progress_bar,
    status_text
):

    downloads_path = str(Path.home() / "Downloads")

    total = len(values)

    for index, value in enumerate(values):

        status_text.text(f"Gerando arquivo: {value}")

        temp_copy = tempfile.NamedTemporaryFile(
            delete=False,
            suffix=".xlsx"
        )

        temp_copy.close()

        shutil.copy(source_file, temp_copy.name)

        workbook = load_workbook(
            temp_copy.name,
            data_only=True
        )

        original_sheet = workbook[sheet_name]

        keep_only_selected_sheet(
            workbook,
            sheet_name
        )

        sheet = workbook[sheet_name]

        preserve_column_widths(
            original_sheet,
            sheet
        )

        preserve_merged_cells(
            original_sheet,
            sheet
        )

        remove_non_matching_rows(
            sheet,
            filter_column,
            value
        )

        remove_excel_tables(sheet)

        convert_formulas_to_values(sheet)

        current_date = datetime.now()

        month = current_date.strftime("%m")
        year = current_date.strftime("%Y")

        safe_value = clean_filename(str(value))

        filename = (
            f"{filter_column}_{safe_value}_{month}_{year}.xlsx"
        )

        final_path = os.path.join(
            downloads_path,
            filename
        )

        workbook.save(final_path)

        os.remove(temp_copy.name)

        progress = int(((index + 1) / total) * 100)

        progress_bar.progress(progress)